"""
Leitor do arquivo Excel com registros para exclusao.
Deteccao flexivel de colunas por substring match (mesmo padrao do projeto pai).
Suporta data como intervalo "dd/mm/yyyy - dd/mm/yyyy".
"""

import logging
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


@dataclass
class RegistroExclusao:
    """Representa uma linha do Excel: cliente + intervalo de vigencia."""
    nome_cliente: str
    data_inicio: str   # dd/mm/yyyy
    data_fim: str       # dd/mm/yyyy


class LeitorExcelExclusao:
    """Le e valida o arquivo Excel de entrada com deteccao flexivel de colunas."""

    # Candidatos ordenados do mais especifico ao mais generico (substring match)
    _CANDIDATOS_NOME = ["NOME DA TABELA", "NOME_CLIENTE", "NOME DO CLIENTE", "NOME CLIENTE", "NOME", "TABELA", "CLIENTE"]
    _CANDIDATOS_DATA = ["DATA VIGENCIA", "DATA_VIGENCIA", "DATA VIGÊNCIA", "VIGENCIA", "VIGÊNCIA", "DATA"]

    def __init__(self, caminho: str | Path, logger: Optional[logging.Logger] = None) -> None:
        self.caminho = Path(caminho)
        self.logger = logger or logging.getLogger("auto_delete_clientes")

    def validar(self) -> None:
        """Verifica se o arquivo existe e tem as colunas necessarias."""
        if not self.caminho.exists():
            raise FileNotFoundError(f"Arquivo Excel nao encontrado: {self.caminho}")

        wb = load_workbook(self.caminho, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        cabecalhos_raw = [str(c.value or "") for c in next(ws.iter_rows(min_row=1, max_row=1))]
        cabecalhos = [self._normalizar(c) for c in cabecalhos_raw]
        wb.close()

        self.logger.info(f"Colunas encontradas no Excel: {cabecalhos_raw}")

        col_nome = self._encontrar_coluna(cabecalhos, self._CANDIDATOS_NOME)
        col_data = self._encontrar_coluna(cabecalhos, self._CANDIDATOS_DATA)

        if col_nome is None:
            raise ValueError(
                f"Coluna de nome nao encontrada. Cabecalhos: {cabecalhos_raw}. "
                f"Esperado uma coluna contendo: {self._CANDIDATOS_NOME}"
            )
        if col_data is None:
            raise ValueError(
                f"Coluna de data nao encontrada. Cabecalhos: {cabecalhos_raw}. "
                f"Esperado uma coluna contendo: {self._CANDIDATOS_DATA}"
            )

        self.logger.info(f"Usando coluna '{cabecalhos_raw[col_nome]}' (indice {col_nome}) como nome")
        self.logger.info(f"Usando coluna '{cabecalhos_raw[col_data]}' (indice {col_data}) como data")

    def ler(self) -> list[RegistroExclusao]:
        """Le todas as linhas validas da primeira aba do Excel."""
        wb = load_workbook(self.caminho, read_only=True, data_only=True)
        ws = wb.worksheets[0]

        linhas = list(ws.iter_rows(min_row=1, values_only=False))
        if not linhas:
            wb.close()
            raise ValueError("Arquivo Excel esta vazio")

        # Detectar colunas
        cabecalhos_raw = [str(c.value or "") for c in linhas[0]]
        cabecalhos = [self._normalizar(c) for c in cabecalhos_raw]

        self.logger.info(f"Colunas encontradas no Excel: {cabecalhos_raw}")

        idx_nome = self._encontrar_coluna(cabecalhos, self._CANDIDATOS_NOME)
        idx_data = self._encontrar_coluna(cabecalhos, self._CANDIDATOS_DATA)

        if idx_nome is None:
            wb.close()
            raise ValueError(f"Coluna de nome nao encontrada. Cabecalhos: {cabecalhos_raw}")
        if idx_data is None:
            wb.close()
            raise ValueError(f"Coluna de data nao encontrada. Cabecalhos: {cabecalhos_raw}")

        self.logger.info(f"Usando coluna '{cabecalhos_raw[idx_nome]}' (indice {idx_nome}) como nome")
        self.logger.info(f"Usando coluna '{cabecalhos_raw[idx_data]}' (indice {idx_data}) como data")

        registros: list[RegistroExclusao] = []
        for numero_linha, linha in enumerate(linhas[1:], start=2):
            nome_raw = linha[idx_nome].value if idx_nome < len(linha) else None
            data_raw = linha[idx_data].value if idx_data < len(linha) else None

            if not nome_raw:
                continue

            nome = self._limpar_texto(nome_raw)
            vigencia = self._limpar_texto(data_raw)

            if not vigencia:
                self.logger.warning(f"Linha {numero_linha}: data vazia para '{nome}', pulando")
                continue

            try:
                data_inicio, data_fim = self._parsear_vigencia(vigencia)
            except ValueError as erro:
                self.logger.warning(f"Linha {numero_linha}: {erro} para '{nome}', pulando")
                continue

            registros.append(RegistroExclusao(
                nome_cliente=nome,
                data_inicio=data_inicio,
                data_fim=data_fim,
            ))

        wb.close()
        self.logger.info(f"Excel lido com sucesso: {len(registros)} registros encontrados")
        return registros

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _encontrar_coluna(self, cabecalhos: list[str], candidatos: list[str]) -> Optional[int]:
        """Substring match: retorna indice da primeira coluna que contem um candidato.
        Usa o mesmo padrao do projeto pai (candidato in nome_coluna).
        """
        candidatos_normalizados = [self._normalizar(c) for c in candidatos]
        for candidato in candidatos_normalizados:
            for i, cab in enumerate(cabecalhos):
                if candidato in cab:
                    return i
        return None

    def _parsear_vigencia(self, vigencia: str) -> tuple[str, str]:
        """Parseia 'dd/mm/yyyy - dd/mm/yyyy' em (data_inicio, data_fim).
        Mesmo padrao do projeto pai (src/servicos/leitor_excel.py).
        """
        if " - " in vigencia:
            inicio, fim = vigencia.split(" - ", 1)
            return inicio.strip(), fim.strip()

        texto = self._normalizar(vigencia).lower()
        for separador in [" a ", " ate "]:
            if separador in texto:
                inicio, fim = texto.split(separador, 1)
                return inicio.strip(), fim.strip()

        raise ValueError(f"Formato de vigencia nao reconhecido: '{vigencia}'")

    def _normalizar(self, texto: str) -> str:
        """Remove acentos, converte para maiusculo e limpa espacos."""
        texto_str = str(texto or "").strip()
        texto_nfkd = unicodedata.normalize("NFKD", texto_str)
        sem_acento = "".join(c for c in texto_nfkd if not unicodedata.combining(c))
        return " ".join(sem_acento.upper().split())

    def _limpar_texto(self, valor) -> str:
        """Converte valor da celula para string limpa."""
        if valor is None:
            return ""
        return str(valor).strip()


def detectar_arquivo_excel(pasta_data: Path) -> Path:
    """Auto-detecta o primeiro .xlsx na pasta data/.
    Retorna o caminho do arquivo encontrado ou levanta erro.
    """
    if not pasta_data.exists():
        raise FileNotFoundError(f"Pasta de dados nao encontrada: {pasta_data}")

    arquivos = sorted(pasta_data.glob("*.xlsx"))
    if not arquivos:
        raise FileNotFoundError(f"Nenhum arquivo .xlsx encontrado em: {pasta_data}")

    return arquivos[0]
