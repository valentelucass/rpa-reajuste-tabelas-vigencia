"""
Script de reprocessamento: le reprocessar.xlsx (gerado pelo main.py)
e tenta excluir novamente os registros que falharam.

Gera um novo reprocessar.xlsx apenas com os que falharam novamente.

Uso:
    cd auto_delete_clientes
    python reprocessar.py
    python reprocessar.py --headless
"""

import sys
from pathlib import Path

import argparse

if __package__ in (None, ""):
    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from auto_delete_compat import carregar_modulo_local

    config = carregar_modulo_local("config")
    from acoes_navegador import AcoesNavegador
    from logger_config import configurar_logger
    from navegador import FabricaNavegador
    from pagina_exclusao import NavegadorFechadoError, PaginaExclusao
    from pagina_login import PaginaLogin
    from utils.atraso_humano import atraso_humano
    from main import RegistroErro, salvar_reprocessamento
else:
    from . import config
    from .acoes_navegador import AcoesNavegador
    from .logger_config import configurar_logger
    from .navegador import FabricaNavegador
    from .pagina_exclusao import NavegadorFechadoError, PaginaExclusao
    from .pagina_login import PaginaLogin
    from .utils.atraso_humano import atraso_humano
    from .main import RegistroErro, salvar_reprocessamento


def ler_reprocessar(caminho: Path, logger) -> list[dict]:
    """Le o arquivo reprocessar.xlsx e retorna lista de registros."""
    from openpyxl import load_workbook

    if not caminho.exists():
        raise FileNotFoundError(f"Arquivo de reprocessamento nao encontrado: {caminho}")

    wb = load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.worksheets[0]

    linhas = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if len(linhas) < 2:
        return []

    registros = []
    for linha in linhas[1:]:
        nome = str(linha[0] or "").strip()
        data = str(linha[1] or "").strip()
        if not nome:
            continue

        # Parsear data_inicio e data_fim do campo "dd/mm/yyyy - dd/mm/yyyy"
        data_inicio, data_fim = "", ""
        if " - " in data:
            partes = data.split(" - ", 1)
            data_inicio = partes[0].strip()
            data_fim = partes[1].strip()

        registros.append({
            "nome_cliente": nome,
            "data_vigencia": data,
            "data_inicio": data_inicio,
            "data_fim": data_fim,
        })

    logger.info(f"Reprocessar: {len(registros)} registros lidos de {caminho.name}")
    return registros


def main() -> None:
    parser = argparse.ArgumentParser(description="Reprocessamento de exclusoes que falharam")
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Executar o navegador em modo headless",
    )
    args = parser.parse_args()

    if args.headless:
        config.HEADLESS = True

    # Logger (novo log a cada execucao)
    logger = configurar_logger()
    logger.info("=" * 60)
    logger.info("INICIO - Reprocessamento de Exclusoes")
    logger.info("=" * 60)

    # Ler reprocessar.xlsx
    caminho_reprocessar = config.ARQUIVO_REPROCESSAMENTO
    try:
        registros = ler_reprocessar(caminho_reprocessar, logger)
    except FileNotFoundError as erro:
        logger.error(str(erro))
        sys.exit(1)

    if not registros:
        logger.info("Nenhum registro para reprocessar. Encerrando.")
        sys.exit(0)

    total = len(registros)
    logger.info(f"Total de registros para reprocessar: {total}")

    # Rastreamento
    resultados = {"sucesso": 0, "erro": 0}
    erros_novos: list[RegistroErro] = []

    driver = None
    try:
        logger.info("Iniciando navegador...")
        driver = FabricaNavegador.criar()
        acoes = AcoesNavegador(driver, logger)

        # Login
        pagina_login = PaginaLogin(acoes, logger)
        pagina_login.abrir()
        pagina_login.autenticar()

        # Navegar para tabelas de cliente
        pagina = PaginaExclusao(acoes, logger)
        pagina.acessar_tabelas_cliente()

        # Configurar filtros uma unica vez
        primeiro = registros[0]
        if primeiro["data_inicio"] and primeiro["data_fim"]:
            pagina.configurar_filtros_iniciais(
                primeiro["data_inicio"], primeiro["data_fim"]
            )

        # Processar cada registro
        for i, reg in enumerate(registros, 1):
            try:
                pagina.verificar_navegador_aberto()
            except NavegadorFechadoError:
                for r in registros[i - 1:]:
                    erros_novos.append(RegistroErro(
                        nome_cliente=r["nome_cliente"],
                        data_vigencia=r["data_vigencia"],
                        motivo="Navegador fechado",
                    ))
                    resultados["erro"] += 1
                break

            logger.info("-" * 40)
            logger.info(f"[{i}/{total}] Buscando cliente: {reg['nome_cliente']}")

            try:
                resultado = pagina.excluir_registro(reg["nome_cliente"])

                if resultado == "sucesso":
                    resultados["sucesso"] += 1
                elif resultado == "ja_processado":
                    resultados["sucesso"] += 1
                elif resultado == "nao_encontrado":
                    resultados["sucesso"] += 1
                    logger.info(
                        f"[{i}/{total}] Cliente '{reg['nome_cliente']}' nao encontrado; "
                        "considerando como ja excluido"
                    )
                else:
                    resultados["erro"] += 1
                    erros_novos.append(RegistroErro(
                        nome_cliente=reg["nome_cliente"],
                        data_vigencia=reg["data_vigencia"],
                        motivo="Erro ao excluir",
                    ))

            except NavegadorFechadoError:
                resultados["erro"] += 1
                erros_novos.append(RegistroErro(
                    nome_cliente=reg["nome_cliente"],
                    data_vigencia=reg["data_vigencia"],
                    motivo="Navegador fechado",
                ))
                break
            except Exception as erro:
                resultados["erro"] += 1
                logger.error(f"[{i}/{total}] ERRO: {reg['nome_cliente']} - {erro}")
                erros_novos.append(RegistroErro(
                    nome_cliente=reg["nome_cliente"],
                    data_vigencia=reg["data_vigencia"],
                    motivo="Erro ao excluir",
                ))
            finally:
                try:
                    pagina.limpar_campo_nome()
                except Exception:
                    pass
                atraso_humano()

        # Sobrescrever reprocessar.xlsx (so com os que falharam de novo)
        salvar_reprocessamento(erros_novos, caminho_reprocessar, logger)

        # Resumo
        logger.info("=" * 60)
        logger.info("Processamento finalizado")
        logger.info(f"  Total processados: {total}")
        logger.info(f"  Total sucesso: {resultados['sucesso']}")
        logger.info(f"  Total erro: {resultados['erro']}")
        logger.info("=" * 60)

    except NavegadorFechadoError:
        salvar_reprocessamento(erros_novos, caminho_reprocessar, logger)
    except Exception as erro:
        logger.critical(f"Erro critico: {erro}")
        salvar_reprocessamento(erros_novos, caminho_reprocessar, logger)
        sys.exit(1)
    finally:
        if driver:
            try:
                logger.info("Fechando navegador...")
                driver.quit()
            except Exception:
                pass

    logger.info("Reprocessamento finalizado.")


if __name__ == "__main__":
    main()
