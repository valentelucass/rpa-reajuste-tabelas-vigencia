"""
Leitura e parsing do arquivo Excel de entrada.
Aba 1: nome da tabela, data de vigencia, percentual.
Aba 2: componentes de reajuste (aba + nome da taxa).
"""

from dataclasses import dataclass
from pathlib import Path
from typing import Optional
import unicodedata

import openpyxl

import config


@dataclass
class DadosTabelaExcel:
    """Dados de uma linha da Aba 1 do Excel."""

    nome: str
    data_inicio: str
    data_fim: str
    percentual: float


@dataclass
class ComponenteReajuste:
    """Dados de uma linha da Aba 2 do Excel."""

    aba: str
    nome_taxa: str


class LeitorExcel:
    """Le e valida o Excel recebido no inicio da operacao."""

    def __init__(self, caminho_excel: str | Path) -> None:
        self.caminho = Path(caminho_excel)
        self._workbook: Optional[openpyxl.Workbook] = None

    def _abrir(self) -> openpyxl.Workbook:
        if self._workbook is None:
            if not self.caminho.exists():
                raise FileNotFoundError(f"Arquivo Excel nao encontrado: {self.caminho}")
            self._workbook = openpyxl.load_workbook(self.caminho, data_only=True)
        return self._workbook

    def validar(self) -> None:
        """Valida a estrutura minima do Excel antes de iniciar a automacao."""
        wb = self._abrir()
        if len(wb.sheetnames) < 2:
            raise ValueError("O Excel deve ter pelo menos 2 abas (Aba 1 e Aba 2).")

        tabelas = self.ler_aba_um()
        if not tabelas:
            raise ValueError("A Aba 1 nao contem linhas de dados.")

        componentes = self.ler_aba_dois()
        if not componentes:
            raise ValueError("A Aba 2 nao contem componentes de reajuste.")

    def ler_aba_um(self) -> list[DadosTabelaExcel]:
        """
        Le a primeira aba do Excel.
        Colunas esperadas (em qualquer ordem): DATA VIGENCIA, NOME DA TABELA, PERCENTUAL.
        """
        wb = self._abrir()
        planilha = wb.worksheets[0]
        cabecalho = self._ler_cabecalho(planilha)

        col_nome = self._encontrar_coluna(cabecalho, ["NOME DA TABELA", "NOME", "TABELA"])
        col_vigencia = self._encontrar_coluna(cabecalho, ["DATA VIGENCIA", "VIGENCIA"])
        col_percentual = self._encontrar_coluna(cabecalho, ["PERCENTUAL", "%", "REAJUSTE"])

        resultado: list[DadosTabelaExcel] = []
        for linha in planilha.iter_rows(min_row=2):
            valores = [celula.value for celula in linha]
            if not any(valores):
                continue

            nome = self._limpar_texto(linha[col_nome].value)
            vigencia = self._limpar_texto(linha[col_vigencia].value)
            percentual = self._parsear_percentual_celula(linha[col_percentual])

            if not nome or not vigencia:
                continue

            data_inicio, data_fim = self._parsear_vigencia(vigencia)
            resultado.append(
                DadosTabelaExcel(
                    nome=nome,
                    data_inicio=data_inicio,
                    data_fim=data_fim,
                    percentual=percentual,
                )
            )

        return resultado

    def ler_aba_dois(self) -> list[ComponenteReajuste]:
        """
        Le a segunda aba do Excel.
        Colunas esperadas: ABA, NOME DA TAXA.
        """
        wb = self._abrir()
        planilha = wb.worksheets[1]
        cabecalho = self._ler_cabecalho(planilha)

        col_aba = self._encontrar_coluna(cabecalho, ["ABA", "TIPO"])
        col_taxa = self._encontrar_coluna(cabecalho, ["NOME DA TAXA", "TAXA", "NOME"])

        resultado: list[ComponenteReajuste] = []
        for linha in planilha.iter_rows(min_row=2, values_only=True):
            if not any(linha):
                continue

            aba_texto = self._limpar_texto(linha[col_aba])
            nome_taxa = self._limpar_texto(linha[col_taxa])
            if not aba_texto or not nome_taxa:
                continue

            aba_id = self._mapear_aba(aba_texto)
            resultado.append(ComponenteReajuste(aba=aba_id, nome_taxa=nome_taxa))

        return resultado

    def _ler_cabecalho(self, planilha) -> list[str]:
        primeira_linha = next(planilha.iter_rows(min_row=1, max_row=1, values_only=True))
        return [self._normalizar_cabecalho(c) for c in primeira_linha]

    def _encontrar_coluna(self, cabecalho: list[str], candidatos: list[str]) -> int:
        candidatos_normalizados = [self._normalizar_cabecalho(candidato) for candidato in candidatos]
        for candidato in candidatos_normalizados:
            for i, nome_coluna in enumerate(cabecalho):
                if candidato in nome_coluna:
                    return i
        raise ValueError(
            f"Nenhuma das colunas {candidatos} encontrada. Colunas disponiveis: {cabecalho}"
        )

    def _normalizar_cabecalho(self, valor) -> str:
        texto = str(valor or "").strip().upper()
        texto_normalizado = unicodedata.normalize("NFKD", texto)
        return texto_normalizado.encode("ASCII", "ignore").decode("ASCII")

    def _limpar_texto(self, valor) -> str:
        if valor is None:
            return ""
        return str(valor).strip()

    def _parsear_vigencia(self, vigencia: str) -> tuple[str, str]:
        """
        Parseia "01/04/2026 - 31/03/2027" em ("01/04/2026", "31/03/2027").
        """
        if " - " in vigencia:
            inicio, fim = vigencia.split(" - ", 1)
            return inicio.strip(), fim.strip()

        texto = self._normalizar_cabecalho(vigencia).lower()
        for separador in [" a ", " ate "]:
            if separador in texto:
                inicio, fim = texto.split(separador, 1)
                return inicio.strip(), fim.strip()

        raise ValueError(f"Formato de DATA VIGENCIA nao reconhecido: '{vigencia}'")

    def _parsear_percentual(self, texto: str) -> float:
        """
        Normaliza percentual para float.
        "9,80%" -> 9.8 | "9.8%" -> 9.8 | "9,8" -> 9.8 | "9.8" -> 9.8
        """
        valor = texto.replace("%", "").replace(",", ".").strip()
        try:
            return float(valor)
        except ValueError:
            return 0.0

    def _parsear_percentual_celula(self, celula) -> float:
        """
        Interpreta texto livre e numeros vindos de celulas formatadas como porcentagem.
        """
        valor = celula.value
        if valor is None:
            return 0.0

        if isinstance(valor, (int, float)):
            numero = float(valor)
            formato = str(getattr(celula, "number_format", "") or "")
            if "%" in formato:
                return numero * 100
            return numero

        return self._parsear_percentual(str(valor))

    def _mapear_aba(self, aba_texto: str) -> str:
        """Mapeia o texto da Aba para o id HTML correspondente."""
        return config.MAPA_ABAS_REAJUSTE.get(
            aba_texto.strip(),
            config.MAPA_ABAS_REAJUSTE.get(aba_texto.strip().title(), "fee"),
        )
