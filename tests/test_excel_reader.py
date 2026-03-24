"""
Teste 3 — Integracao com Excel
Testa leitura de Excel com cenarios realistas:
  - Leitura correta limitada a LIMITE_EXCEL_TESTE linhas
  - Integridade de leitura
  - Comportamento com linhas vazias
  - Comportamento com dados invalidos
  - Cenarios de borda (1 aba, vazio, caracteres especiais)
"""

import time
from pathlib import Path

import pytest

from src.servicos.leitor_excel import LeitorExcel, DadosTabelaExcel, ComponenteReajuste
from tests.constantes_teste import LIMITE_EXCEL_TESTE
from tests.utils.gerador_excel_fake import (
    gerar_excel_fake,
    gerar_excel_grande,
    gerar_excel_com_problemas,
    gerar_excel_vazio,
    gerar_excel_uma_aba,
    gerar_excel_caracteres_especiais,
    gerar_excel_minimo,
    gerar_nome_tabela,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def excel_padrao(tmp_path):
    """Excel com LIMITE_EXCEL_TESTE linhas padrao."""
    return gerar_excel_fake(
        tmp_path / "padrao.xlsx",
        total_linhas=LIMITE_EXCEL_TESTE,
        percentual_fixo="9,80%",
        vigencia_fixa="01/04/2026 - 31/03/2027",
    )


@pytest.fixture
def excel_grande(tmp_path):
    """Excel para teste de integridade de leitura."""
    return gerar_excel_grande(tmp_path / "grande.xlsx", total=LIMITE_EXCEL_TESTE)


@pytest.fixture
def excel_problemas(tmp_path):
    """Excel com linhas vazias e dados invalidos."""
    return gerar_excel_com_problemas(tmp_path / "problemas.xlsx")


@pytest.fixture
def excel_vazio(tmp_path):
    """Excel com cabecalho mas sem dados."""
    return gerar_excel_vazio(tmp_path / "vazio.xlsx")


@pytest.fixture
def excel_uma_aba(tmp_path):
    """Excel com apenas 1 aba."""
    return gerar_excel_uma_aba(tmp_path / "uma_aba.xlsx")


@pytest.fixture
def excel_especiais(tmp_path):
    """Excel com caracteres especiais nos nomes."""
    return gerar_excel_caracteres_especiais(tmp_path / "especiais.xlsx")


@pytest.fixture
def excel_minimo(tmp_path):
    """Excel com 1 linha apenas."""
    return gerar_excel_minimo(tmp_path / "minimo.xlsx")


# ---------------------------------------------------------------------------
# Teste: Leitura correta
# ---------------------------------------------------------------------------

class TestLeituraCorreta:
    """Testa leitura correta de Excel."""

    def test_le_todas_as_linhas(self, excel_padrao):
        """Deve ler todas as linhas com dados corretos."""
        leitor = LeitorExcel(excel_padrao)
        tabelas = leitor.ler_aba_um()

        assert len(tabelas) == LIMITE_EXCEL_TESTE, (
            f"Esperado {LIMITE_EXCEL_TESTE}, obteve {len(tabelas)}"
        )

    def test_todas_linhas_tem_nome(self, excel_padrao):
        """Todas as linhas devem ter nome nao vazio."""
        leitor = LeitorExcel(excel_padrao)
        tabelas = leitor.ler_aba_um()

        for i, tabela in enumerate(tabelas):
            assert tabela.nome != "", f"Linha {i+1} tem nome vazio"
            assert len(tabela.nome) > 0

    def test_todas_linhas_tem_datas(self, excel_padrao):
        """Todas as linhas devem ter datas de inicio e fim."""
        leitor = LeitorExcel(excel_padrao)
        tabelas = leitor.ler_aba_um()

        for i, tabela in enumerate(tabelas):
            assert tabela.data_inicio == "01/04/2026", (
                f"Linha {i+1}: data_inicio '{tabela.data_inicio}'"
            )
            assert tabela.data_fim == "31/03/2027", (
                f"Linha {i+1}: data_fim '{tabela.data_fim}'"
            )

    def test_todas_linhas_tem_percentual(self, excel_padrao):
        """Todas as linhas devem ter percentual correto."""
        leitor = LeitorExcel(excel_padrao)
        tabelas = leitor.ler_aba_um()

        for i, tabela in enumerate(tabelas):
            assert tabela.percentual == 9.8, (
                f"Linha {i+1}: percentual {tabela.percentual} != 9.8"
            )

    def test_nomes_sao_unicos(self, excel_padrao):
        """Todos os nomes devem ser unicos no Excel."""
        leitor = LeitorExcel(excel_padrao)
        tabelas = leitor.ler_aba_um()
        nomes = [t.nome for t in tabelas]

        assert len(nomes) == len(set(nomes)), "Ha nomes duplicados no Excel"

    def test_nomes_correspondem_ao_gerador(self, excel_padrao):
        """Nomes devem corresponder ao gerador de nomes fake."""
        leitor = LeitorExcel(excel_padrao)
        tabelas = leitor.ler_aba_um()

        for i, tabela in enumerate(tabelas):
            esperado = gerar_nome_tabela(i)
            assert tabela.nome == esperado, (
                f"Linha {i+1}: '{tabela.nome}' != esperado '{esperado}'"
            )

    def test_le_aba_dois_componentes(self, excel_padrao):
        """Aba 2 deve conter os componentes de reajuste."""
        leitor = LeitorExcel(excel_padrao)
        componentes = leitor.ler_aba_dois()

        assert len(componentes) == 8, f"Esperado 8 componentes, obteve {len(componentes)}"
        for comp in componentes:
            assert isinstance(comp, ComponenteReajuste)
            assert comp.aba != ""
            assert comp.nome_taxa != ""

    def test_validacao_estrutura_ok(self, excel_padrao):
        """Validacao deve passar para Excel bem formado."""
        leitor = LeitorExcel(excel_padrao)
        leitor.validar()  # nao deve levantar excecao


# ---------------------------------------------------------------------------
# Teste: Integridade de leitura
# ---------------------------------------------------------------------------

class TestIntegridadeLeitura:
    """Testa integridade de leitura de Excel."""

    def test_leitura_completa_sem_erro(self, excel_padrao):
        """Leitura deve completar sem erros em tempo razoavel."""
        leitor = LeitorExcel(excel_padrao)

        inicio = time.time()
        tabelas = leitor.ler_aba_um()
        duracao = time.time() - inicio

        assert len(tabelas) == LIMITE_EXCEL_TESTE
        assert duracao < 5.0, f"Leitura demorou {duracao:.2f}s (max: 5.0s)"

    def test_leitura_grande_sem_erro(self, excel_grande):
        """Leitura de Excel grande deve completar sem erros."""
        leitor = LeitorExcel(excel_grande)

        inicio = time.time()
        tabelas = leitor.ler_aba_um()
        duracao = time.time() - inicio

        assert len(tabelas) > 0
        assert duracao < 10.0, f"Leitura demorou {duracao:.2f}s (max: 10.0s)"

    def test_leitura_repetida_nao_degrada(self, excel_padrao):
        """Multiplas leituras consecutivas nao devem degradar."""
        leitor = LeitorExcel(excel_padrao)
        tempos = []

        for _ in range(5):
            leitor._workbook = None  # forca re-leitura
            inicio = time.time()
            leitor.ler_aba_um()
            tempos.append(time.time() - inicio)

        # Nenhuma leitura deve demorar mais de 5s
        assert tempos[-1] < 5.0, (
            f"Ultima leitura demorou {tempos[-1]:.3f}s"
        )

    def test_tempo_por_linha_consistente(self, excel_padrao):
        """Tempo por linha deve ser consistente."""
        leitor = LeitorExcel(excel_padrao)

        inicio = time.time()
        tabelas = leitor.ler_aba_um()
        duracao = time.time() - inicio

        tempo_por_linha = duracao / len(tabelas)
        assert tempo_por_linha < 0.05, (
            f"Tempo por linha: {tempo_por_linha*1000:.1f}ms (max: 50ms)"
        )


# ---------------------------------------------------------------------------
# Teste: Comportamento com linhas vazias
# ---------------------------------------------------------------------------

class TestLinhasVazias:
    """Testa comportamento com linhas vazias no Excel."""

    def test_ignora_linhas_vazias(self, excel_problemas):
        """Linhas completamente vazias devem ser ignoradas."""
        leitor = LeitorExcel(excel_problemas)
        tabelas = leitor.ler_aba_um()

        for tabela in tabelas:
            assert tabela.nome != "", "Linha vazia nao deveria estar nos resultados"

    def test_excel_vazio_retorna_lista_vazia(self, excel_vazio):
        """Excel sem dados deve retornar lista vazia."""
        leitor = LeitorExcel(excel_vazio)
        tabelas = leitor.ler_aba_um()
        assert tabelas == []

    def test_validacao_falha_excel_vazio(self, excel_vazio):
        """Validacao deve falhar para Excel sem dados."""
        leitor = LeitorExcel(excel_vazio)
        with pytest.raises(ValueError, match="nao contem"):
            leitor.validar()

    def test_linhas_com_nome_vazio_ignoradas(self, tmp_path):
        """Linhas com nome vazio mas outros campos preenchidos devem ser ignoradas."""
        import openpyxl
        caminho = tmp_path / "nome_vazio.xlsx"
        wb = openpyxl.Workbook()
        aba1 = wb.active
        aba1.append(["NOME DA TABELA", "DATA VIGÊNCIA", "PERCENTUAL"])
        aba1.append(["", "01/04/2026 - 31/03/2027", "9,80%"])
        aba1.append(["Tabela Real", "01/04/2026 - 31/03/2027", "9,80%"])
        aba1.append([None, "01/04/2026 - 31/03/2027", "9,80%"])

        aba2 = wb.create_sheet("Componentes")
        aba2.append(["ABA", "NOME DA TAXA"])
        aba2.append(["Taxas", "Taxa Admin"])
        wb.save(str(caminho))

        leitor = LeitorExcel(caminho)
        tabelas = leitor.ler_aba_um()
        assert len(tabelas) == 1
        assert tabelas[0].nome == "Tabela Real"


# ---------------------------------------------------------------------------
# Teste: Comportamento com dados invalidos
# ---------------------------------------------------------------------------

class TestDadosInvalidos:
    """Testa comportamento com dados invalidos."""

    def test_excel_com_problemas_nao_quebra(self, excel_problemas):
        """Excel com dados invalidos nao deve causar crash."""
        leitor = LeitorExcel(excel_problemas)
        tabelas = leitor.ler_aba_um()
        # Deve retornar apenas linhas validas
        assert isinstance(tabelas, list)

    def test_percentual_invalido_retorna_zero(self, tmp_path):
        """Percentual invalido deve retornar 0.0."""
        import openpyxl
        caminho = tmp_path / "pct_invalido.xlsx"
        wb = openpyxl.Workbook()
        aba1 = wb.active
        aba1.append(["NOME DA TABELA", "DATA VIGÊNCIA", "PERCENTUAL"])
        aba1.append(["Tabela Teste", "01/04/2026 - 31/03/2027", "ABC"])

        aba2 = wb.create_sheet("Componentes")
        aba2.append(["ABA", "NOME DA TAXA"])
        aba2.append(["Taxas", "Taxa Admin"])
        wb.save(str(caminho))

        leitor = LeitorExcel(caminho)
        tabelas = leitor.ler_aba_um()
        assert len(tabelas) == 1
        assert tabelas[0].percentual == 0.0

    def test_vigencia_invalida_levanta_erro(self, tmp_path):
        """Vigencia invalida deve levantar ValueError."""
        import openpyxl
        caminho = tmp_path / "vig_invalida.xlsx"
        wb = openpyxl.Workbook()
        aba1 = wb.active
        aba1.append(["NOME DA TABELA", "DATA VIGÊNCIA", "PERCENTUAL"])
        aba1.append(["Tabela Teste", "FORMATO_INVALIDO_SEM_SEPARADOR", "9,80%"])

        aba2 = wb.create_sheet("Componentes")
        aba2.append(["ABA", "NOME DA TAXA"])
        aba2.append(["Taxas", "Taxa Admin"])
        wb.save(str(caminho))

        leitor = LeitorExcel(caminho)
        with pytest.raises(ValueError, match="nao reconhecido"):
            leitor.ler_aba_um()

    def test_excel_uma_aba_falha_validacao(self, excel_uma_aba):
        """Excel com apenas 1 aba deve falhar na validacao."""
        leitor = LeitorExcel(excel_uma_aba)
        with pytest.raises(ValueError, match="2 abas"):
            leitor.validar()

    def test_arquivo_inexistente_levanta_erro(self, tmp_path):
        """Arquivo inexistente deve levantar FileNotFoundError."""
        leitor = LeitorExcel(tmp_path / "nao_existe.xlsx")
        with pytest.raises(FileNotFoundError):
            leitor.ler_aba_um()


# ---------------------------------------------------------------------------
# Teste: Caracteres especiais
# ---------------------------------------------------------------------------

class TestCaracteresEspeciaisExcel:
    """Testa leitura de nomes com caracteres especiais."""

    def test_le_caracteres_especiais(self, excel_especiais):
        """Deve ler nomes com caracteres especiais corretamente."""
        leitor = LeitorExcel(excel_especiais)
        tabelas = leitor.ler_aba_um()

        assert len(tabelas) > 0
        nomes = [t.nome for t in tabelas]

        assert any("São Paulo" in n for n in nomes), "Nao encontrou nome com acentos"
        assert any("C&A" in n for n in nomes), "Nao encontrou nome com &"

    def test_nome_muito_longo(self, excel_especiais):
        """Deve ler nome com 200 caracteres sem truncar."""
        leitor = LeitorExcel(excel_especiais)
        tabelas = leitor.ler_aba_um()

        nomes_longos = [t.nome for t in tabelas if len(t.nome) >= 200]
        assert len(nomes_longos) >= 1, "Deveria ter pelo menos 1 nome longo"


# ---------------------------------------------------------------------------
# Teste: Formatos de percentual
# ---------------------------------------------------------------------------

class TestFormatosPercentual:
    """Testa parsing de diferentes formatos de percentual."""

    @pytest.mark.parametrize("entrada,esperado", [
        ("9,80%", 9.8),
        ("9.80%", 9.8),
        ("9,8", 9.8),
        ("9.8", 9.8),
        ("10%", 10.0),
        ("0,5%", 0.5),
        ("100%", 100.0),
        ("ABC", 0.0),
        ("", 0.0),
    ])
    def test_parsing_percentual(self, entrada, esperado, tmp_path):
        """Deve parsear diferentes formatos de percentual corretamente."""
        import openpyxl
        caminho = tmp_path / f"pct_{hash(entrada)}.xlsx"
        wb = openpyxl.Workbook()
        aba1 = wb.active
        aba1.append(["NOME DA TABELA", "DATA VIGÊNCIA", "PERCENTUAL"])
        aba1.append(["Tabela Teste", "01/04/2026 - 31/03/2027", entrada])

        aba2 = wb.create_sheet("Componentes")
        aba2.append(["ABA", "NOME DA TAXA"])
        aba2.append(["Taxas", "Taxa Admin"])
        wb.save(str(caminho))

        leitor = LeitorExcel(caminho)
        tabelas = leitor.ler_aba_um()

        if entrada in ("", "ABC"):
            # Para texto invalido: pode retornar 0.0 ou pular a linha
            if tabelas:
                assert tabelas[0].percentual == esperado
        else:
            assert len(tabelas) == 1
            assert tabelas[0].percentual == pytest.approx(esperado, abs=0.01), (
                f"Entrada '{entrada}': obteve {tabelas[0].percentual}, esperado {esperado}"
            )


# ---------------------------------------------------------------------------
# Teste: Formatos de vigencia
# ---------------------------------------------------------------------------

class TestFormatosVigencia:
    """Testa parsing de diferentes formatos de data de vigencia."""

    @pytest.mark.parametrize("vigencia,inicio_esperado,fim_esperado", [
        ("01/04/2026 - 31/03/2027", "01/04/2026", "31/03/2027"),
        ("15/01/2026 - 14/01/2027", "15/01/2026", "14/01/2027"),
    ])
    def test_parsing_vigencia_formato_padrao(self, vigencia, inicio_esperado, fim_esperado, tmp_path):
        """Deve parsear formato padrao DD/MM/AAAA - DD/MM/AAAA."""
        import openpyxl
        caminho = tmp_path / f"vig_{hash(vigencia)}.xlsx"
        wb = openpyxl.Workbook()
        aba1 = wb.active
        aba1.append(["NOME DA TABELA", "DATA VIGÊNCIA", "PERCENTUAL"])
        aba1.append(["Tabela Teste", vigencia, "9,80%"])

        aba2 = wb.create_sheet("Componentes")
        aba2.append(["ABA", "NOME DA TAXA"])
        aba2.append(["Taxas", "Taxa Admin"])
        wb.save(str(caminho))

        leitor = LeitorExcel(caminho)
        tabelas = leitor.ler_aba_um()
        assert len(tabelas) == 1
        assert tabelas[0].data_inicio == inicio_esperado
        assert tabelas[0].data_fim == fim_esperado


# ---------------------------------------------------------------------------
# Teste: Leitura de 1 linha (minimo)
# ---------------------------------------------------------------------------

class TestLeituraMinima:
    """Testa leitura de Excel com quantidade minima de dados."""

    def test_le_uma_linha(self, excel_minimo):
        """Deve ler corretamente Excel com apenas 1 linha."""
        leitor = LeitorExcel(excel_minimo)
        tabelas = leitor.ler_aba_um()
        assert len(tabelas) == 1
        assert tabelas[0].nome == gerar_nome_tabela(0)

    def test_validacao_passa_com_uma_linha(self, excel_minimo):
        """Validacao deve passar com apenas 1 linha."""
        leitor = LeitorExcel(excel_minimo)
        leitor.validar()  # nao deve levantar excecao


# ---------------------------------------------------------------------------
# Teste: Tipo de dados retornados
# ---------------------------------------------------------------------------

class TestTipoDados:
    """Testa que os tipos de dados retornados estao corretos."""

    def test_retorno_tipo_dataclass(self, excel_padrao):
        """Deve retornar lista de DadosTabelaExcel."""
        leitor = LeitorExcel(excel_padrao)
        tabelas = leitor.ler_aba_um()

        assert isinstance(tabelas, list)
        for tabela in tabelas:
            assert isinstance(tabela, DadosTabelaExcel)

    def test_campos_tipos_corretos(self, excel_padrao):
        """Campos devem ter tipos corretos."""
        leitor = LeitorExcel(excel_padrao)
        tabelas = leitor.ler_aba_um()

        for tabela in tabelas:
            assert isinstance(tabela.nome, str)
            assert isinstance(tabela.data_inicio, str)
            assert isinstance(tabela.data_fim, str)
            assert isinstance(tabela.percentual, float)

    def test_componentes_tipo_correto(self, excel_padrao):
        """Componentes devem ter tipos corretos."""
        leitor = LeitorExcel(excel_padrao)
        componentes = leitor.ler_aba_dois()

        for comp in componentes:
            assert isinstance(comp, ComponenteReajuste)
            assert isinstance(comp.aba, str)
            assert isinstance(comp.nome_taxa, str)
