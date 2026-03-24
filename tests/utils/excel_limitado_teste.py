"""Helpers para criar uma copia temporaria do Excel com volume reduzido."""

from pathlib import Path
from tempfile import gettempdir
from uuid import uuid4

import openpyxl

from tests.constantes_teste import LIMITE_EXCEL_TESTE


def criar_excel_limitado(
    origem: str | Path,
    limite: int = LIMITE_EXCEL_TESTE,
) -> Path:
    """
    Cria uma copia temporaria do Excel limitando a primeira aba ao numero
    maximo de registros permitido para testes.
    """
    caminho_origem = Path(origem)
    workbook_origem = openpyxl.load_workbook(caminho_origem)
    workbook_destino = openpyxl.Workbook()

    for indice_aba, aba_origem in enumerate(workbook_origem.worksheets):
        aba_destino = (
            workbook_destino.active
            if indice_aba == 0
            else workbook_destino.create_sheet()
        )
        aba_destino.title = aba_origem.title

        linhas = list(aba_origem.iter_rows(values_only=True))
        if not linhas:
            continue

        if indice_aba == 0:
            for linha in linhas[: limite + 1]:
                aba_destino.append(list(linha))
            continue

        for linha in linhas:
            aba_destino.append(list(linha))

    caminho_destino = Path(gettempdir()) / f"rpa_teste_limitado_{uuid4().hex[:8]}.xlsx"
    workbook_destino.save(caminho_destino)
    return caminho_destino
