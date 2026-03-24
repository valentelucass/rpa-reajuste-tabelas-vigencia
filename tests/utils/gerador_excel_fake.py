"""
Gerador de arquivos Excel falsos para testes.
Gera massa de dados limitada conforme constantes de teste.
"""

import random
import string
from pathlib import Path
from typing import Optional

import openpyxl

from tests.constantes_teste import LIMITE_EXCEL_TESTE, MODO_TESTE


NOMES_BASE = [
    "Tabela Transportes Norte", "Tabela Logistica Sul", "Tabela Frete Expresso",
    "Tabela Cargas Especiais", "Tabela Regional Sudeste", "Tabela Interestadual",
    "Tabela Rodoviario Premium", "Tabela Aereo Nacional", "Tabela Maritimo",
    "Tabela Carga Pesada", "Tabela Encomendas", "Tabela Mudancas",
    "Tabela Frigorificado", "Tabela Carga Viva", "Tabela Granel Solido",
    "Tabela Container", "Tabela Porta a Porta", "Tabela Expressa 24h",
    "Tabela Last Mile", "Tabela Cross Docking",
]

COMPONENTES_ABA = [
    ("Taxas", "Taxa de Administracao"),
    ("Taxas", "Taxa de Seguro"),
    ("Taxas", "Taxa de Manuseio"),
    ("Excedentes", "Excesso de Peso"),
    ("Excedentes", "Excesso de Volume"),
    ("Adicionais", "Seguro Adicional"),
    ("Adicionais", "Coleta Especial"),
    ("Adicionais", "Entrega Agendada"),
]


def gerar_nome_tabela(indice: int) -> str:
    """Gera um nome de tabela unico baseado no indice."""
    base = NOMES_BASE[indice % len(NOMES_BASE)]
    sufixo = indice // len(NOMES_BASE)
    if sufixo == 0:
        return base
    return f"{base} {sufixo:03d}"


def gerar_percentual() -> str:
    """Gera percentual aleatorio no formato brasileiro."""
    valor = round(random.uniform(1.0, 15.0), 2)
    return f"{valor}".replace(".", ",") + "%"


def gerar_vigencia(ano: int = 2026) -> str:
    """Gera data de vigencia no formato DD/MM/AAAA - DD/MM/AAAA."""
    mes_inicio = random.randint(1, 6)
    dia_inicio = random.randint(1, 28)
    mes_fim = mes_inicio + random.randint(3, 6)
    if mes_fim > 12:
        mes_fim = 12
    dia_fim = random.randint(1, 28)
    return f"{dia_inicio:02d}/{mes_inicio:02d}/{ano} - {dia_fim:02d}/{mes_fim:02d}/{ano}"


def gerar_excel_fake(
    caminho: Path,
    total_linhas: int = LIMITE_EXCEL_TESTE,
    incluir_linhas_vazias: bool = False,
    incluir_dados_invalidos: bool = False,
    percentual_fixo: Optional[str] = None,
    vigencia_fixa: Optional[str] = None,
) -> Path:
    """
    Gera um arquivo Excel fake com a estrutura esperada pelo LeitorExcel.

    Args:
        caminho: Caminho do arquivo .xlsx a ser gerado.
        total_linhas: Numero de linhas de dados na Aba 1.
        incluir_linhas_vazias: Se True, insere linhas vazias aleatorias.
        incluir_dados_invalidos: Se True, insere dados malformados.
        percentual_fixo: Se informado, usa este percentual em todas as linhas.
        vigencia_fixa: Se informado, usa esta vigencia em todas as linhas.

    Returns:
        Path do arquivo gerado.
    """
    if MODO_TESTE and total_linhas > LIMITE_EXCEL_TESTE:
        total_linhas = LIMITE_EXCEL_TESTE

    wb = openpyxl.Workbook()

    # --- Aba 1: Tabelas ---
    aba1 = wb.active
    aba1.title = "Tabelas"
    aba1.append(["NOME DA TABELA", "DATA VIGÊNCIA", "PERCENTUAL"])

    linhas_inseridas = 0
    indice_nome = 0

    while linhas_inseridas < total_linhas:
        # Linhas vazias aleatorias
        if incluir_linhas_vazias and random.random() < 0.03:
            aba1.append([None, None, None])
            linhas_inseridas += 1
            continue

        # Dados invalidos aleatorios (apenas cenarios que o LeitorExcel ignora silenciosamente)
        if incluir_dados_invalidos and random.random() < 0.02:
            cenario = random.choice(["nome_vazio", "percentual_invalido"])
            if cenario == "nome_vazio":
                aba1.append(["", gerar_vigencia(), gerar_percentual()])
            else:
                aba1.append([gerar_nome_tabela(indice_nome), gerar_vigencia(), "ABC"])
                indice_nome += 1
            linhas_inseridas += 1
            continue

        nome = gerar_nome_tabela(indice_nome)
        vigencia = vigencia_fixa or gerar_vigencia()
        percentual = percentual_fixo or gerar_percentual()

        aba1.append([nome, vigencia, percentual])
        indice_nome += 1
        linhas_inseridas += 1

    # --- Aba 2: Componentes de Reajuste ---
    aba2 = wb.create_sheet("Componentes")
    aba2.append(["ABA", "NOME DA TAXA"])

    for aba_nome, taxa_nome in COMPONENTES_ABA:
        aba2.append([aba_nome, taxa_nome])

    wb.save(str(caminho))
    return caminho


def gerar_excel_minimo(caminho: Path) -> Path:
    """Gera Excel minimo com 1 linha valida para testes rapidos."""
    return gerar_excel_fake(caminho, total_linhas=1)


def gerar_excel_grande(caminho: Path, total: int = LIMITE_EXCEL_TESTE) -> Path:
    """Gera Excel para testes de integridade de leitura."""
    return gerar_excel_fake(caminho, total_linhas=total)


def gerar_excel_com_problemas(caminho: Path) -> Path:
    """Gera Excel deterministico com 1 linha valida, 1 vazia e 1 invalida."""
    wb = openpyxl.Workbook()
    aba1 = wb.active
    aba1.title = "Tabelas"
    aba1.append(["NOME DA TABELA", "DATA VIGÊNCIA", "PERCENTUAL"])
    aba1.append(["Tabela Valida", "01/04/2026 - 31/03/2027", "9,80%"])
    aba1.append([None, None, None])
    aba1.append(["", "01/04/2026 - 31/03/2027", "ABC"])

    aba2 = wb.create_sheet("Componentes")
    aba2.append(["ABA", "NOME DA TAXA"])
    for aba_nome, taxa_nome in COMPONENTES_ABA:
        aba2.append([aba_nome, taxa_nome])

    wb.save(str(caminho))
    return caminho


def gerar_excel_vazio(caminho: Path) -> Path:
    """Gera Excel com cabecalho mas sem dados."""
    wb = openpyxl.Workbook()
    aba1 = wb.active
    aba1.title = "Tabelas"
    aba1.append(["NOME DA TABELA", "DATA VIGÊNCIA", "PERCENTUAL"])

    aba2 = wb.create_sheet("Componentes")
    aba2.append(["ABA", "NOME DA TAXA"])

    wb.save(str(caminho))
    return caminho


def gerar_excel_uma_aba(caminho: Path) -> Path:
    """Gera Excel com apenas 1 aba (deve falhar na validacao)."""
    wb = openpyxl.Workbook()
    aba1 = wb.active
    aba1.title = "Tabelas"
    aba1.append(["NOME DA TABELA", "DATA VIGÊNCIA", "PERCENTUAL"])
    aba1.append(["Tabela Teste", "01/01/2026 - 31/12/2026", "9,80%"])
    wb.save(str(caminho))
    return caminho


def gerar_excel_caracteres_especiais(caminho: Path) -> Path:
    """Gera Excel com nomes contendo caracteres especiais (limitado a 3)."""
    wb = openpyxl.Workbook()
    aba1 = wb.active
    aba1.title = "Tabelas"
    aba1.append(["NOME DA TABELA", "DATA VIGÊNCIA", "PERCENTUAL"])

    nomes_especiais = [
        "Tabela São Paulo — Especial",
        "Tabela C&A / Logística",
        "A" * 200,  # nome muito longo
    ]
    for nome in nomes_especiais:
        aba1.append([nome, "01/04/2026 - 31/03/2027", "9,80%"])

    aba2 = wb.create_sheet("Componentes")
    aba2.append(["ABA", "NOME DA TAXA"])
    aba2.append(["Taxas", "Taxa de Administracao"])

    wb.save(str(caminho))
    return caminho
